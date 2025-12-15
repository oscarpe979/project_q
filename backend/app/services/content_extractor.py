"""
Universal Content Extractor for CD Grid files (PDF and Excel).
Extracts raw grid data preserving structure, merge info, and cell positions.
"""
from typing import Dict, Any, List, BinaryIO, Union
import io
import datetime


class ContentExtractor:
    """Extract raw grid data from PDF/Excel regardless of format."""
    
    def extract(self, file_obj: Union[str, BinaryIO], filename: str) -> Dict[str, Any]:
        """
        Extract raw cell data from file.
        
        Returns standardized format:
        {
            "type": "excel" | "pdf",
            "cells": [{"row": 1, "col": 1, "value": "text"}, ...],
            "merges": [{"start_row": 1, "end_row": 2, "start_col": 1, "end_col": 3}, ...],
            "dimensions": {"rows": 50, "cols": 10}
        }
        """
        filename_lower = filename.lower()
        
        if filename_lower.endswith(('.xlsx', '.xls')):
            return self._extract_excel(file_obj, filename_lower)
        elif filename_lower.endswith('.pdf'):
            return self._extract_pdf(file_obj)
        else:
            raise ValueError(f"Unsupported file type: {filename}")
    
    def _extract_excel(self, file_obj: Union[str, BinaryIO], filename: str) -> Dict[str, Any]:
        """Extract Excel data preserving merge information."""
        from openpyxl import load_workbook
        import pandas as pd
        
        # Handle .xls files by converting to .xlsx format first
        if filename.endswith('.xls'):
            return self._extract_legacy_excel(file_obj)
        
        # Load workbook (data_only=True to get values instead of formulas)
        if isinstance(file_obj, str):
            wb = load_workbook(file_obj, data_only=True)
        else:
            wb = load_workbook(file_obj, data_only=True)
        
        ws = wb.active
        cells = []
        
        # Extract all non-empty cells with position
        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            for col_idx, cell in enumerate(row, start=1):
                if cell.value is not None:
                    # Clean and format the value
                    val = cell.value
                    value = ""
                    
                    # Handle Excel Date/Time objects
                    if isinstance(val, datetime.time):
                        # Format "17:00:00" -> "5:00 pm"
                        value = val.strftime("%I:%M %p").lstrip("0").lower()
                        
                    elif isinstance(val, datetime.datetime):
                        # If it's a pure date (midnight): Keep YYYY-MM-DD for headers
                        if val.time() == datetime.time(0, 0, 0):
                             value = val.strftime("%Y-%m-%d")
                        else:
                             # It has a time component, treat as Time (e.g. "1900-01-01 17:00:00")
                             value = val.strftime("%I:%M %p").lstrip("0").lower()
                             
                    else:
                        value = str(val).strip()
                        
                    if value:
                        cells.append({
                            "row": row_idx,
                            "col": col_idx,
                            "value": value,
                            "col_letter": cell.column_letter
                        })
        
        # Extract merge ranges
        merges = []
        for merged_range in ws.merged_cells.ranges:
            merges.append({
                "start_row": merged_range.min_row,
                "end_row": merged_range.max_row,
                "start_col": merged_range.min_col,
                "end_col": merged_range.max_col
            })
        
        return {
            "type": "excel",
            "cells": cells,
            "merges": merges,
            "dimensions": {
                "rows": ws.max_row or 0,
                "cols": ws.max_column or 0
            }
        }
    
    def _extract_legacy_excel(self, file_obj: Union[str, BinaryIO]) -> Dict[str, Any]:
        """Extract .xls files using pandas (limited merge support)."""
        import pandas as pd
        
        if isinstance(file_obj, str):
            df = pd.read_excel(file_obj, header=None)
        else:
            df = pd.read_excel(file_obj, header=None)
            if hasattr(file_obj, 'seek'):
                file_obj.seek(0)
        
        cells = []
        for row_idx in range(len(df)):
            for col_idx in range(len(df.columns)):
                value = df.iloc[row_idx, col_idx]
                if pd.notna(value):
                    str_value = str(value).strip()
                    if str_value:
                        cells.append({
                            "row": row_idx + 1,
                            "col": col_idx + 1,
                            "value": str_value
                        })
        
        return {
            "type": "excel",
            "cells": cells,
            "merges": [],  # Legacy format has limited merge support
            "dimensions": {
                "rows": len(df),
                "cols": len(df.columns)
            }
        }
    
    def _extract_pdf(self, file_obj: Union[str, BinaryIO]) -> Dict[str, Any]:
        """Extract PDF table data using pdfplumber's default detection."""
        import pdfplumber
        
        # Read PDF bytes
        if isinstance(file_obj, str):
            with open(file_obj, 'rb') as f:
                pdf_bytes = f.read()
        else:
            pdf_bytes = file_obj.read()
            if hasattr(file_obj, 'seek'):
                file_obj.seek(0)
        
        cells = []
        max_row = 0
        max_col = 0
        
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page_num, page in enumerate(pdf.pages):
                # Use default table extraction - works best for most PDFs
                tables = page.extract_tables()
                
                for table in tables:
                    if not table:
                        continue
                    
                    for row_idx, row in enumerate(table):
                        if not row:
                            continue
                        for col_idx, cell_text in enumerate(row):
                            if cell_text is not None:
                                value = str(cell_text).strip()
                                if value:
                                    cells.append({
                                        "row": row_idx + 1,
                                        "col": col_idx + 1,
                                        "value": value,
                                        "page": page_num + 1
                                    })
                                    max_row = max(max_row, row_idx + 1)
                                    max_col = max(max_col, col_idx + 1)
        
        return {
            "type": "pdf",
            "cells": cells,
            "merges": [],
            "dimensions": {
                "rows": max_row,
                "cols": max_col
            }
        }
    
    def format_for_llm(self, data: Dict[str, Any], max_cells: int = 150) -> str:
        """
        Format extracted data as a string for LLM consumption.
        Limits output to max_cells to control token usage.
        """
        lines = []
        lines.append(f"Document type: {data['type'].upper()}")
        lines.append(f"Dimensions: {data['dimensions']['rows']} rows x {data['dimensions']['cols']} columns")
        
        if data['merges']:
            lines.append(f"Merged cell ranges: {len(data['merges'])}")
        
        lines.append("")
        lines.append("Cell data (row, col: value):")
        lines.append("-" * 40)
        
        # Sort cells by row then column for logical order
        sorted_cells = sorted(data['cells'], key=lambda c: (c['row'], c['col']))
        
        for i, cell in enumerate(sorted_cells[:max_cells]):
            lines.append(f"Row {cell['row']}, Col {cell['col']}: \"{cell['value']}\"")
        
        if len(sorted_cells) > max_cells:
            lines.append(f"... and {len(sorted_cells) - max_cells} more cells")
        
        return "\n".join(lines)
