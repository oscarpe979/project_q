from fastapi import APIRouter, Depends, HTTPException, status
from sqlmodel import Session, select
from typing import List, Optional
from datetime import datetime, timezone, time, timedelta
import io
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Helper to parse a single time string with various formats
def parse_single_time_string(t_str: Optional[str]) -> Optional[time]:
    if not t_str:
        return None
    
    t_str = t_str.strip().lower().replace("midnight", "12:00 am")
    formats = [
        "%I:%M %p",  # 07:00 am
        "%I:%M%p",   # 07:00am
        "%H:%M",     # 19:00
        "%-I:%M %p", # 7:00 am
    ]
    
    for fmt in formats:
        try:
            return datetime.strptime(t_str, fmt).time()
        except ValueError:
            continue
    return None

# Helper to parse time string "7:00 am - 4:30 pm"
def parse_port_times(time_str: Optional[str]):
    if not time_str:
        return None, None
    
    # Normalize string: lowercase, remove extra spaces
    ts = time_str.lower().strip()
    
    # Handle "Midnight" special case (often used for departure)
    if "midnight" in ts:
        # If it's a range like "7:00 am - midnight"
        parts = ts.replace('midnight', '12:00 am').split('-')
    else:
        # Split by various dash types
        parts = ts.replace('–', '-').replace('—', '-').split('-')
        
    if len(parts) != 2:
        return None, None
    
    arrival = parse_single_time_string(parts[0])
    departure = parse_single_time_string(parts[1])
    
    return arrival, departure



from backend.app.db.session import get_session
from backend.app.db.models import (
    User, Voyage, VoyageItinerary, ScheduleItem, Venue, EventType, VenueHighlight, VenueSchedule
)
from backend.app.api.v1.endpoints.auth import get_current_user
from backend.app.schemas.schedules import (
    EventInput, ItineraryInput, OtherVenueShowInput, PublishScheduleRequest
)

router = APIRouter(
    prefix="/schedules",
    tags=["schedules"],
    responses={404: {"description": "Not found"}},
)

# --- Endpoints ---

@router.post("/", status_code=status.HTTP_201_CREATED)
def publish_schedule(
    request: PublishScheduleRequest,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    if not current_user.venue_id:
        raise HTTPException(status_code=400, detail="User must be assigned to a venue to publish a schedule.")
    
    if not current_user.ship_id:
        raise HTTPException(status_code=400, detail="User must be assigned to a ship to publish a schedule.")

    # 1. Find or Create Voyage
    # First, checked if the target voyage number already exists
    existing_voyage = session.exec(select(Voyage).where(Voyage.voyage_number == request.voyage_number)).first()

    # STRICT VALIDATION Logic
    if request.original_voyage_number:
        # Update Mode: We are editing an existing schedule (or renaming it)
        # Security Check: ideally we should check if original_voyage_number exists, but for now we trust the flow
        # Conflict Check: If we are renaming (target != original), target must NOT exist.
        if request.voyage_number != request.original_voyage_number:
            if existing_voyage:
                 raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT, 
                    detail=f"Cannot overwrite existing schedule '{request.voyage_number}' while renaming from '{request.original_voyage_number}'."
                )
    else:
        # Creation Mode: We are creating a NEW draft/schedule
        # Target must NOT exist.
        if existing_voyage:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT, 
                detail=f"Voyage '{request.voyage_number}' already exists."
            )
            
    voyage = existing_voyage
    
    # Calculate start/end dates from itinerary if available
    start_date = datetime.now().date()
    end_date = datetime.now().date()
    
    if request.itinerary:
        try:
            sorted_itinerary = sorted(request.itinerary, key=lambda x: x.day)
            # Try to parse start/end dates
            try:
                start_date = datetime.strptime(sorted_itinerary[0].date, "%Y-%m-%d").date()
            except ValueError:
                start_date = datetime.now().date()
            
            try:
                end_date = datetime.strptime(sorted_itinerary[-1].date, "%Y-%m-%d").date()
            except ValueError:
                # Fallback: use start_date + days
                end_date = start_date + timedelta(days=sorted_itinerary[-1].day - 1)
        except Exception:
            pass # Keep defaults if anything fails

    if not voyage:
        voyage = Voyage(
            ship_id=current_user.ship_id,
            voyage_number=request.voyage_number,
            start_date=start_date,
            end_date=end_date,
            created_at=datetime.now(),
            updated_at=datetime.now()
        )
        session.add(voyage)
        session.commit()
        session.refresh(voyage)
    else:
        voyage.updated_at = datetime.now()
        # Update dates if itinerary was provided
        if request.itinerary:
            voyage.start_date = start_date
            voyage.end_date = end_date
            
        session.add(voyage)
        session.commit()

    # Ensure VenueSchedule exists (for both new and existing voyages)
    venue_schedule = session.exec(
        select(VenueSchedule)
        .where(VenueSchedule.venue_id == current_user.venue_id)
        .where(VenueSchedule.voyage_id == voyage.id)
    ).first()
    
    if not venue_schedule:
        venue_schedule = VenueSchedule(
            venue_id=current_user.venue_id,
            voyage_id=voyage.id
        )
        session.add(venue_schedule)
    else:
        venue_schedule.updated_at = datetime.now()
        session.add(venue_schedule)
        
    # 2. Update Itinerary (Shared)
    # Strategy: Delete existing for this voyage and re-insert. 
    # NOTE: This might affect other venues if they rely on the same itinerary. 
    # For now, we assume the itinerary is shared and the latest publish wins/updates it.
    # To be safe, we could check if itinerary items exist and only update if different, 
    # but deleting and re-inserting is cleaner for "publishing" the source of truth.
    
    # Check if we should update itinerary. Only if provided.
    if request.itinerary:
        existing_itineraries = session.exec(select(VoyageItinerary).where(VoyageItinerary.voyage_id == voyage.id)).all()
        for item in existing_itineraries:
            session.delete(item)
        
        for item in request.itinerary:
            if item.arrival or item.departure:
                # Use provided structured times
                arrival_time = parse_single_time_string(item.arrival)
                departure_time = parse_single_time_string(item.departure)
            else:
                # Fallback to parsing the 'time' string
                arrival_time, departure_time = parse_port_times(item.time)
            
            try:
                itinerary_date = datetime.strptime(item.date, "%Y-%m-%d").date()
            except ValueError:
                # Fallback for empty or invalid dates: Use today + day offset
                itinerary_date = datetime.now().date() + timedelta(days=item.day - 1)

            itinerary_entry = VoyageItinerary(
                voyage_id=voyage.id,
                day_number=item.day,
                date=itinerary_date,
                location=item.location.title(),
                arrival_time=arrival_time,
                departure_time=departure_time
            )
            session.add(itinerary_entry)

    # 3. Update Schedule Items (Scoped to Venue)
    # Delete existing items for this venue and voyage
    existing_items = session.exec(
        select(ScheduleItem)
        .where(ScheduleItem.voyage_id == voyage.id)
        .where(ScheduleItem.venue_id == current_user.venue_id)
    ).all()
    
    for item in existing_items:
        session.delete(item)
        
    # Insert new items
    for event in request.events:
        # Find or create EventType? 
        # For MVP, we might just store the type string if we don't strictly enforce EventType table yet
        # But ScheduleItem requires event_type_id (Optional).
        # Let's try to find an EventType by name, if not found, maybe leave null or create?
        # Model: event_type_id is Optional.
        
        # We need to map frontend 'type' to backend 'type' field in ScheduleItem
        
        new_item = ScheduleItem(
            voyage_id=voyage.id,
            venue_id=current_user.venue_id,
            title=event.title,
            start_time=event.start,
            end_time=event.end,
            type=event.type,
            time_display=event.time_display,
            notes=event.notes
        )
        session.add(new_item)
        
    # 4. Update Venue Highlights (Scoped to Venue)
    if request.other_venue_shows is not None:
        # Delete existing highlights for this venue and voyage
        existing_highlights = session.exec(
            select(VenueHighlight)
            .where(VenueHighlight.voyage_id == voyage.id)
            .where(VenueHighlight.source_venue_id == current_user.venue_id)
        ).all()
        
        for h in existing_highlights:
            session.delete(h)
            
        # Add new highlights
        for show in request.other_venue_shows:
            try:
                highlight_date = datetime.strptime(show.date, "%Y-%m-%d").date()
            except ValueError:
                # Skip invalid dates for highlights as we can't easily infer them
                continue

            highlight = VenueHighlight(
                voyage_id=voyage.id,
                source_venue_id=current_user.venue_id,
                date=highlight_date,
                highlight_venue_name=show.venue,
                title=show.title,
                time_text=show.time
            )
            session.add(highlight)

    session.commit()
    
    return {"message": "Schedule published successfully", "voyage_number": voyage.voyage_number}

@router.get("/latest")
def get_latest_schedule(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    if not current_user.venue_id:
        raise HTTPException(status_code=400, detail="User must be assigned to a venue.")
    
    if not current_user.ship_id:
        raise HTTPException(status_code=400, detail="User must be assigned to a ship.")

    # Find most recent voyage for this ship AND venue
    # We look for the ScheduleItem with the highest ID for this venue, 
    # as items are re-inserted on publish (so higher ID = later publish).
    latest_item_id = session.exec(
        select(ScheduleItem.voyage_id)
        .where(ScheduleItem.venue_id == current_user.venue_id)
        .order_by(ScheduleItem.id.desc())
    ).first()
    
    if not latest_item_id:
        return {"events": [], "itinerary": []}
        
    voyage = session.exec(select(Voyage).where(Voyage.id == latest_item_id)).first()
    
    if not voyage:
        return {"events": [], "itinerary": []}

    # Get Itinerary
    itinerary_items = session.exec(
        select(VoyageItinerary)
        .where(VoyageItinerary.voyage_id == voyage.id)
        .order_by(VoyageItinerary.day_number)
    ).all()
    
    # Get Schedule Items for user's venue
    schedule_items = session.exec(
        select(ScheduleItem)
        .where(ScheduleItem.voyage_id == voyage.id)
        .where(ScheduleItem.venue_id == current_user.venue_id)
        .where(ScheduleItem.venue_id == current_user.venue_id)
    ).all()
    
    # Get Venue Highlights
    highlights = session.exec(
        select(VenueHighlight)
        .where(VenueHighlight.voyage_id == voyage.id)
        .where(VenueHighlight.source_venue_id == current_user.venue_id)
    ).all()
    
    # Format response
    formatted_itinerary = [
        {
            "day": item.day_number,
            "date": item.date.isoformat(),
            "location": item.location,
            "arrival_time": item.arrival_time.strftime('%-I:%M %p').lower() if item.arrival_time else None,
            "departure_time": item.departure_time.strftime('%-I:%M %p').lower() if item.departure_time else None
        }
        for item in itinerary_items
    ]
    
    formatted_events = [
        {
            "title": item.title,
            "start": item.start_time.isoformat(),
            "end": item.end_time.isoformat(),
            "type": item.type,
            "time_display": item.time_display,
            "notes": item.notes
        }
        for item in schedule_items
    ]
    
    formatted_highlights = [
        {
            "venue": h.highlight_venue_name,
            "date": h.date.isoformat(),
            "title": h.title,
            "time": h.time_text
        }
        for h in highlights
    ]

    return {
        "voyage_number": voyage.voyage_number,
        "events": formatted_events,
        "itinerary": formatted_itinerary,
        "other_venue_shows": formatted_highlights
    }

@router.get("/")
def list_schedules(
    search: Optional[str] = None,
    skip: int = 0,
    limit: int = 20,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    if not current_user.venue_id:
        raise HTTPException(status_code=400, detail="User must be assigned to a venue.")
    
    from backend.app.services.search import SearchService
    search_service = SearchService(session)
    
    voyages = search_service.search_schedules(search, current_user.venue_id, skip=skip, limit=limit)
    
    return [
        {
            "voyage_number": v.voyage_number,
            "start_date": v.start_date.isoformat(),
            "end_date": v.end_date.isoformat(),
        }
        for v in voyages
    ]

@router.get("/venues")
def get_ship_venues(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    if not current_user.ship_id:
        raise HTTPException(status_code=400, detail="User must be assigned to a ship.")
        
    venues = session.exec(select(Venue).where(Venue.ship_id == current_user.ship_id)).all()
    
    # Filter out the current user's venue if desired, or keep all. 
    # The user wants "Other Venue Shows", so typically we exclude the current one.
    # But for a "template", maybe they want to see all? 
    # Let's exclude the current venue to match the "Other" semantics.
    other_venues = [v for v in venues if v.id != current_user.venue_id]
    
    return [{"id": v.id, "name": v.name} for v in other_venues]

@router.get("/{voyage_number}")
def get_schedule_by_voyage(
    voyage_number: str,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    if not current_user.venue_id:
        raise HTTPException(status_code=400, detail="User must be assigned to a venue.")
        
    voyage = session.exec(select(Voyage).where(Voyage.voyage_number == voyage_number)).first()
    if not voyage:
        raise HTTPException(status_code=404, detail="Voyage not found")
        
    # Get Itinerary
    itinerary_items = session.exec(
        select(VoyageItinerary)
        .where(VoyageItinerary.voyage_id == voyage.id)
        .order_by(VoyageItinerary.day_number)
    ).all()
    
    # Get Schedule Items for user's venue
    schedule_items = session.exec(
        select(ScheduleItem)
        .where(ScheduleItem.voyage_id == voyage.id)
        .where(ScheduleItem.venue_id == current_user.venue_id)
    ).all()
    
    # Get Venue Highlights
    highlights = session.exec(
        select(VenueHighlight)
        .where(VenueHighlight.voyage_id == voyage.id)
        .where(VenueHighlight.source_venue_id == current_user.venue_id)
    ).all()
    
    # Format response
    formatted_itinerary = [
        {
            "day": item.day_number,
            "date": item.date.isoformat(),
            "location": item.location,
            "arrival_time": item.arrival_time.strftime('%-I:%M %p').lower() if item.arrival_time else None,
            "departure_time": item.departure_time.strftime('%-I:%M %p').lower() if item.departure_time else None
        }
        for item in itinerary_items
    ]
    
    formatted_events = [
        {
            "title": item.title,
            "start": item.start_time.isoformat(),
            "end": item.end_time.isoformat(),
            "type": item.type,
            "time_display": item.time_display,
            "notes": item.notes
        }
        for item in schedule_items
    ]
    
    formatted_highlights = [
        {
            "venue": h.highlight_venue_name,
            "date": h.date.isoformat(),
            "title": h.title,
            "time": h.time_text
        }
        for h in highlights
    ]
    
    return {
        "voyage_number": voyage.voyage_number,
        "events": formatted_events,
        "itinerary": formatted_itinerary,
        "other_venue_shows": formatted_highlights
    }



@router.delete("/{voyage_number}")
def delete_schedule(
    voyage_number: str,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    if not current_user.venue_id:
        raise HTTPException(status_code=400, detail="User must be assigned to a venue.")
        
    voyage = session.exec(select(Voyage).where(Voyage.voyage_number == voyage_number)).first()
    if not voyage:
        raise HTTPException(status_code=404, detail="Voyage not found")
        
    # Delete Venue Highlights for this venue (and voyage)
    highlights_to_delete = session.exec(
        select(VenueHighlight)
        .where(VenueHighlight.voyage_id == voyage.id)
        .where(VenueHighlight.source_venue_id == current_user.venue_id)
    ).all()
    
    for highlight in highlights_to_delete:
        session.delete(highlight)

    # Delete Schedule Items for this venue
    items_to_delete = session.exec(
        select(ScheduleItem)
        .where(ScheduleItem.voyage_id == voyage.id)
        .where(ScheduleItem.venue_id == current_user.venue_id)
    ).all()
    
    count = len(items_to_delete)
    for item in items_to_delete:
        session.delete(item)
        
    # Delete VenueSchedule for this venue
    venue_schedule = session.exec(
        select(VenueSchedule)
        .where(VenueSchedule.venue_id == current_user.venue_id)
        .where(VenueSchedule.voyage_id == voyage.id)
    ).first()
    
    if venue_schedule:
        session.delete(venue_schedule)

    session.commit() # Commit deletion of items and schedule first
    
    # Check if any VenueSchedule remains for this voyage (any venue)
    remaining_schedules = session.exec(
        select(VenueSchedule).where(VenueSchedule.voyage_id == voyage.id)
    ).all()
    
    if not remaining_schedules:
        # No venues are using this voyage anymore, delete the Voyage and its Itinerary
        itineraries = session.exec(
            select(VoyageItinerary).where(VoyageItinerary.voyage_id == voyage.id)
        ).all()
        for it in itineraries:
            session.delete(it)
            
        # Also delete ALL remaining highlights for this voyage (from any venue)
        all_highlights = session.exec(
            select(VenueHighlight).where(VenueHighlight.voyage_id == voyage.id)
        ).all()
        for h in all_highlights:
            session.delete(h)
            
        session.delete(voyage)
        session.commit()
        return {"message": f"Deleted schedule and the Voyage {voyage_number} as it is now unused."}
    
    return {"message": f"Deleted schedule for voyage {voyage_number}"}

@router.get("/{voyage_number}/export")
def export_schedule(
    voyage_number: str,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    if not current_user.venue_id:
        raise HTTPException(status_code=400, detail="User must be assigned to a venue.")

    # 1. Fetch Data
    voyage = session.exec(select(Voyage).where(Voyage.voyage_number == voyage_number)).first()
    if not voyage:
        raise HTTPException(status_code=404, detail="Voyage not found")

    venue = session.exec(select(Venue).where(Venue.id == current_user.venue_id)).first()
    venue_name = venue.name if venue else "Venue"
    ship_code = current_user.username.split('_')[0].upper() if current_user.username else "SHIP"

    itinerary_items = session.exec(
        select(VoyageItinerary)
        .where(VoyageItinerary.voyage_id == voyage.id)
        .order_by(VoyageItinerary.day_number)
    ).all()

    schedule_items = session.exec(
        select(ScheduleItem)
        .where(ScheduleItem.voyage_id == voyage.id)
        .where(ScheduleItem.venue_id == current_user.venue_id)
    ).all()

    # Fetch Venue Highlights for Footer
    highlights = session.exec(
        select(VenueHighlight)
        .where(VenueHighlight.voyage_id == voyage.id)
        .where(VenueHighlight.source_venue_id == current_user.venue_id)
    ).all()

    # 2. Prepare Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    # User requested sheet name to be the venue name
    # Excel sheet names have a max length of 31 chars and cannot contain certain chars.
    # We should sanitize it briefly, but for now assuming venue_name is safe enough or short enough.
    safe_title = venue_name[:31].replace(':', '').replace('\\', '').replace('/', '').replace('?', '').replace('*', '').replace('[', '').replace(']', '')
    ws.title = safe_title
    
    # Set default zoom to 84%
    ws.sheet_view.zoomScale = 84

    # Styles
    # Styles
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    medium_side = Side(style='medium')
    header_border = Border(left=medium_side, right=medium_side, top=medium_side, bottom=medium_side)
    grid_border = Border(left=medium_side, right=medium_side, top=Side(style='thin'), bottom=Side(style='thin'))
    
    bold_font = Font(name='Arial', size=11, bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # 3. Build Grid Structure
    # Row 1: Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(itinerary_items) + 2)
    ws.cell(row=1, column=1, value=f"{venue_name} Schedule - VY {voyage_number}").font = Font(name='Arial', size=28, bold=True)
    ws.cell(row=1, column=1).alignment = center_align
    ws.row_dimensions[1].height = 37.5 # 50 pixels
    # Apply header border to title row
    for c in range(1, len(itinerary_items) + 3):
        ws.cell(row=1, column=c).border = header_border

    # Headers
    headers = ["DAY", "DATE", "", "PORT", ""] # Row labels
    # Row 2: DAY numbers
    # Col 1 & Last: Font 10
    ws.cell(row=2, column=1, value="DAY").font = Font(name='Arial', size=10, bold=True)
    ws.cell(row=2, column=1).alignment = center_align
    ws.cell(row=2, column=1).border = header_border
    ws.cell(row=2, column=len(itinerary_items) + 2, value="DAY").font = Font(name='Arial', size=10, bold=True)
    ws.cell(row=2, column=len(itinerary_items) + 2).alignment = center_align
    ws.cell(row=2, column=len(itinerary_items) + 2).border = header_border

    # Row 3: Day Names (DATE label)
    # Merge Row 3 & 4 for DATE label
    ws.merge_cells(start_row=3, start_column=1, end_row=4, end_column=1)
    ws.cell(row=3, column=1, value="DATE").font = Font(name='Arial', size=10, bold=True)
    ws.cell(row=3, column=1).alignment = center_align
    ws.cell(row=3, column=1).border = header_border
    # Apply border to merged cell (row 4)
    ws.cell(row=4, column=1).border = header_border

    ws.merge_cells(start_row=3, start_column=len(itinerary_items) + 2, end_row=4, end_column=len(itinerary_items) + 2)
    ws.cell(row=3, column=len(itinerary_items) + 2, value="DATE").font = Font(name='Arial', size=10, bold=True)
    ws.cell(row=3, column=len(itinerary_items) + 2).alignment = center_align
    ws.cell(row=3, column=len(itinerary_items) + 2).border = header_border
    # Apply border to merged cell (row 4)
    ws.cell(row=4, column=len(itinerary_items) + 2).border = header_border
    
    # Row 5: PORT
    # Merge Row 5 & 6 for PORT label
    ws.merge_cells(start_row=5, start_column=1, end_row=6, end_column=1)
    ws.cell(row=5, column=1, value="PORT").font = Font(name='Arial', size=10, bold=True)
    ws.cell(row=5, column=1).alignment = center_align
    ws.cell(row=5, column=1).border = header_border
    # Apply border to merged cell (row 6)
    ws.cell(row=6, column=1).border = header_border

    ws.merge_cells(start_row=5, start_column=len(itinerary_items) + 2, end_row=6, end_column=len(itinerary_items) + 2)
    ws.cell(row=5, column=len(itinerary_items) + 2, value="PORT").font = Font(name='Arial', size=10, bold=True)
    ws.cell(row=5, column=len(itinerary_items) + 2).alignment = center_align
    ws.cell(row=5, column=len(itinerary_items) + 2).border = header_border
    # Apply border to merged cell (row 6)
    ws.cell(row=6, column=len(itinerary_items) + 2).border = header_border

    day_col_border = Border(left=medium_side, right=medium_side)
    white_fill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid') # Yellow
    black_fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid') # Black

    # Fill Itinerary Columns (Cols 2 to N+1)
    for idx, item in enumerate(itinerary_items):
        col_idx = idx + 2
        
        # Day Number (Row 2) - Font Size 14, Yellow Background
        ws.cell(row=2, column=col_idx, value=f"DAY {item.day_number}").alignment = center_align
        ws.cell(row=2, column=col_idx).border = header_border
        ws.cell(row=2, column=col_idx).font = Font(name='Arial', size=14, bold=True)
        ws.cell(row=2, column=col_idx).fill = yellow_fill
        
        # Day Name
        ws.cell(row=3, column=col_idx, value=item.date.strftime("%A")).alignment = center_align
        ws.cell(row=3, column=col_idx).border = header_border
        ws.cell(row=3, column=col_idx).font = Font(name='Arial', size=11, bold=True)

        # Date (Row 4) - Black Background, White Font
        ws.cell(row=4, column=col_idx, value=item.date.strftime("%m/%d/%y")).alignment = center_align
        ws.cell(row=4, column=col_idx).border = header_border
        ws.cell(row=4, column=col_idx).font = Font(name='Arial', size=11, bold=True, color='FFFFFFFF')
        ws.cell(row=4, column=col_idx).fill = black_fill

        # Port Name
        ws.cell(row=5, column=col_idx, value=item.location.upper()).alignment = center_align
        ws.cell(row=5, column=col_idx).border = header_border
        ws.cell(row=5, column=col_idx).font = Font(name='Arial', size=11, bold=True)

        # Port Times
        times_str = ""
        if item.arrival_time and item.departure_time:
            times_str = f"{item.arrival_time.strftime('%-I:%M %p').lower()} - {item.departure_time.strftime('%-I:%M %p').lower()}"
        elif item.arrival_time:
             times_str = f"Arr {item.arrival_time.strftime('%-I:%M %p').lower()}"
        elif item.departure_time:
             times_str = f"Dep {item.departure_time.strftime('%-I:%M %p').lower()}"
        
        ws.cell(row=6, column=col_idx, value=times_str).alignment = center_align
        ws.cell(row=6, column=col_idx).border = header_border
        ws.cell(row=6, column=col_idx).font = Font(name='Arial', size=11, bold=True)

    # 4. Time Slots
    start_time = time(7, 0) # 7:00 AM
    end_time = time(0, 0) # Midnight (next day effectively)
    
    # Generate 15 min slots
    # We'll use a loop from 7:00 AM until we hit midnight or 2AM? 
    # Template shows 8:00am to 12:00pm (midnight) usually.
    # Let's go from 7:00 AM to 1:00 AM next day to be safe, or just 24h?
    # Let's assume 7:00 AM to 1:00 AM
    
    current_dt = datetime.combine(datetime.today(), start_time)
    # Grid finishes at 1:00 AM next day
    end_dt = datetime.combine(datetime.today() + timedelta(days=1), time(1, 0))
    
    row_idx = 7
    time_map = {} # time_str -> row_idx

    day_col_border = Border(left=medium_side, right=medium_side)
    white_fill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')

    while current_dt <= end_dt:
        t_str = current_dt.strftime("%-I:%M%p").lower()
        
        # Left Time Column
        ws.cell(row=row_idx, column=1, value=t_str).alignment = Alignment(horizontal='center', vertical='top')
        ws.cell(row=row_idx, column=1).font = Font(name='Arial', size=10, bold=True)
        
        # Right Time Column
        ws.cell(row=row_idx, column=len(itinerary_items) + 2, value=t_str).alignment = Alignment(horizontal='center', vertical='top')
        ws.cell(row=row_idx, column=len(itinerary_items) + 2).font = Font(name='Arial', size=10, bold=True)
        
        # Add borders to grid cells (including time columns)
        for c in range(1, len(itinerary_items) + 3):
            cell = ws.cell(row=row_idx, column=c)
            if 2 <= c <= len(itinerary_items) + 1:
                # Day Columns: No horizontal gridlines, White background
                cell.border = day_col_border
                cell.fill = white_fill
            else:
                # Time Columns: Keep gridlines
                cell.border = grid_border

        time_map[current_dt.time()] = row_idx
        
        current_dt += timedelta(minutes=15)
        row_idx += 1
    
    max_grid_row = row_idx - 1

    # 5. Place Events
    
    # Color Palette (ARGB for openpyxl)
    COLORS = {
        'PRODUCTION_SHOW_1': 'FF963333', # Vivid Red
        'PRODUCTION_SHOW_2': 'FF7E46BE', # Vivid Blue
        'PRODUCTION_SHOW_3': 'FF820080', # Deep Purple
        'HEADLINER': 'FF84F0E6',         # Bright Teal
        'MOVIE': 'FFE1BEE7',             # Light Purple
        'GAME_SHOW': 'FFF3B344',         # Vivid Orange
        'ACTIVITY': 'FFBBDEFB',          # Light Blues
        'MUSIC': 'FF9BFA9E',             # Bright Green
        'PARTY': 'FFA5E1F8',             # Yellow
        'COMEDY': 'FF4DC7A2',            # Light Pink
        'OTHER': 'FFE3DED3',             # Warm Grey
    }

    # Logic to assign colors to production shows
    production_shows = []
    for event in schedule_items:
        if event.type == 'show' and event.title not in production_shows:
            production_shows.append(event.title)
    
    production_color_map = {}
    for idx, title in enumerate(production_shows):
        if idx == 0:
            production_color_map[title] = COLORS['PRODUCTION_SHOW_1']
        elif idx == 1:
            production_color_map[title] = COLORS['PRODUCTION_SHOW_2']
        else:
            production_color_map[title] = COLORS['PRODUCTION_SHOW_3']

    # Helper for contrast color
    def get_contrast_color(hex_color):
        if not hex_color: return 'FF000000' # Default black
        # Remove alpha if present (ARGB -> RGB)
        hex_color = hex_color.replace('#', '')
        if len(hex_color) == 8:
            hex_color = hex_color[2:] # Strip alpha
        
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)
        
        yiq = ((r * 299) + (g * 587) + (b * 114)) / 1000
        return 'FF000000' if yiq >= 128 else 'FFFFFFFF' # Black or White

    for event in schedule_items:
        # Find column based on date
        # We need to match event date to itinerary date
        # Note: schedule_items store start_time as datetime, so we have the date.
        event_date = event.start_time.date()
        event_time = event.start_time.time()

        # Logic for late night events (00:00 - 04:00)
        # Shift them to the previous day so they appear at the end of that day's column
        if event_time < time(4, 0):
            event_date = event_date - timedelta(days=1)
        
        # Find matching itinerary column
        
        # Find matching itinerary column
        col_idx = -1
        for idx, item in enumerate(itinerary_items):
            if item.date == event_date:
                col_idx = idx + 2
                break
        
        if col_idx == -1:
            continue # Event date not in itinerary?

        # Find row based on time
        # Round down to nearest 15 min
        event_time = event.start_time.time()
        # Simple rounding: remove seconds/microseconds
        # We need to handle if time is not exactly on 15 min mark? 
        # For now assume it snaps or we map to nearest previous slot.
        
        # Normalize to 15 min
        minute = event_time.minute
        rounded_minute = (minute // 15) * 15
        rounded_time = time(event_time.hour, rounded_minute)
        
        if rounded_time in time_map:
            r_idx = time_map[rounded_time]
            
            # Calculate duration in 15 min slots
            duration = event.end_time - event.start_time
            slots = int(duration.total_seconds() / 900) # 900s = 15m
            if slots < 1: slots = 1
            
            # Clamp slots to grid end
            if r_idx + slots - 1 > max_grid_row:
                slots = max_grid_row - r_idx + 1

            # Determine Color
            fill_color = COLORS['OTHER']
            event_type = event.type.lower() if event.type else 'other'
            
            if event_type == 'show':
                fill_color = production_color_map.get(event.title, COLORS['PRODUCTION_SHOW_1'])
            elif event_type == 'headliner':
                fill_color = COLORS['HEADLINER']
            elif event_type == 'movie':
                fill_color = COLORS['MOVIE']
            elif event_type == 'game':
                fill_color = COLORS['GAME_SHOW']
            elif event_type == 'activity':
                fill_color = COLORS['ACTIVITY']
            elif event_type == 'music':
                fill_color = COLORS['MUSIC']
            elif event_type == 'party':
                fill_color = COLORS['PARTY']
            elif event_type == 'comedy':
                fill_color = COLORS['COMEDY']
            
            text_color = get_contrast_color(fill_color)

            # Apply style to all cells in the range (No Merging)
            # Text Placement Logic:
            # If slots > 1:
            #   Row 0: Time Range
            #   Middle Row (of remaining): Title
            # If slots == 1:
            #   Row 0: Title (Time implied by grid or too small to fit both)
            
            if event.time_display:
                time_str = event.time_display
            else:
                start_t_str = event.start_time.strftime('%-I:%M%p').lower()
                end_t_str = event.end_time.strftime('%-I:%M%p').lower()
                
                # Handle "late" text
                # If end time is between 1:00 AM and 7:00 AM, show "late"
                # Note: time comparison works on hour/minute. 
                if time(1, 0) < event.end_time.time() < time(7, 0):
                    end_t_str = "late"
                
                time_str = f"{start_t_str} - {end_t_str}"
            
            # Determine title row offset (relative to event start)
            # User requested centering, specifically for odd rows to be in the middle (or "inner closest to top" which implies middle for 3).
            # slots // 2 gives:
            # 1 -> 0
            # 2 -> 1 (Time at 0, Title at 1)
            # 3 -> 1 (Time at 0, Title at 1, Empty at 2)
            # 4 -> 2 (Empty at 0, Time at 1, Title at 2, Empty at 3)
            # 5 -> 2 (Empty at 0, Time at 1, Title at 2, Empty at 3, Empty at 4)
            title_row_offset = slots // 2

            for i in range(slots):
                current_row = r_idx + i
                cell = ws.cell(row=current_row, column=col_idx)
                
                # Apply Fill
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                
                # Apply Borders for Outline Effect
                # Use medium for left/right to match grid
                top_side = Side(style='thin') if i == 0 else Side(style=None)
                bottom_side = Side(style='thin') if i == slots - 1 else Side(style=None)
                left_side = Side(style='medium')
                right_side = Side(style='medium')
                
                cell.border = Border(left=left_side, right=right_side, top=top_side, bottom=bottom_side)
                
                # Text Content
                cell.value = None
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False) # No wrap to prevent height expansion

                if slots > 1:
                    if i == title_row_offset - 1:
                        cell.value = time_str
                        cell.font = Font(name='Arial', size=11, bold=True, color=text_color) # Smaller font for time
                    elif i == title_row_offset: 
                        cell.value = event.title
                        cell.font = Font(name='Arial', size=11, bold=True, color=text_color) # Reduced size to 10
                else:
                    # Single slot
                    cell.value = event.title
                    cell.font = Font(name='Arial', size=11, bold=True, color=text_color) # Reduced size to 10
    
    # 6. Final Formatting
    # Freeze Panes (Rows 1-6)
    ws.freeze_panes = 'A7'

    # Column widths
    # User requested 75px for time columns and 300px for day columns.
    # OpenPyXL uses character units. Approx 1 char = 7px.
    # 75px / 7 ≈ 10.7 -> 12 (generous) -> User requested 100px -> 14.3
    # 300px / 7 ≈ 42.8 -> 42
    ws.column_dimensions[get_column_letter(1)].width = 14.3
    ws.column_dimensions[get_column_letter(len(itinerary_items) + 2)].width = 14.3
    for i in range(len(itinerary_items)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 42

    # Thick Outside Border (Medium as requested)
    thick_side = Side(style='medium')
    
    # Top Edge (Row 1 - Title)
    # Note: Row 1 is merged from col 1 to last_col
    # We need to apply top border to all cells in the range, or just the main merged cell?
    # Applying to all cells in the range is safer for openpyxl.
    for c in range(1, len(itinerary_items) + 3):
        current_border = ws.cell(row=1, column=c).border
        ws.cell(row=1, column=c).border = Border(top=thick_side, left=current_border.left, right=current_border.right, bottom=current_border.bottom)
    
    # Bottom Edge (max_grid_row)
    for c in range(1, len(itinerary_items) + 3):
        # We need to preserve existing borders (thin) and add thick bottom
        current_border = ws.cell(row=max_grid_row, column=c).border
        ws.cell(row=max_grid_row, column=c).border = Border(top=current_border.top, left=current_border.left, right=current_border.right, bottom=thick_side)

    # Left Edge (Col 1)
    for r in range(1, max_grid_row + 1):
        current_border = ws.cell(row=r, column=1).border
        ws.cell(row=r, column=1).border = Border(top=current_border.top, left=thick_side, right=current_border.right, bottom=current_border.bottom)

    # Right Edge (Last Col)
    last_col = len(itinerary_items) + 2
    for r in range(1, max_grid_row + 1):
        current_border = ws.cell(row=r, column=last_col).border
        ws.cell(row=r, column=last_col).border = Border(top=current_border.top, left=current_border.left, right=thick_side, bottom=current_border.bottom)
    
    # Fix corners
    # Top-Left
    ws.cell(row=1, column=1).border = Border(top=thick_side, left=thick_side, right=ws.cell(row=1, column=1).border.right, bottom=ws.cell(row=1, column=1).border.bottom)
    # Top-Right
    ws.cell(row=1, column=last_col).border = Border(top=thick_side, left=ws.cell(row=1, column=last_col).border.left, right=thick_side, bottom=ws.cell(row=1, column=last_col).border.bottom)
    # Bottom-Left
    ws.cell(row=max_grid_row, column=1).border = Border(top=ws.cell(row=max_grid_row, column=1).border.top, left=thick_side, right=ws.cell(row=max_grid_row, column=1).border.right, bottom=thick_side)
    # Bottom-Right
    ws.cell(row=max_grid_row, column=last_col).border = Border(top=ws.cell(row=max_grid_row, column=last_col).border.top, left=ws.cell(row=max_grid_row, column=last_col).border.left, right=thick_side, bottom=thick_side)

    # 7. Other Venue Shows Footer
    if highlights:
        # Group highlights by venue
        venue_shows = {}
        for h in highlights:
            if h.highlight_venue_name not in venue_shows:
                venue_shows[h.highlight_venue_name] = {}
            venue_shows[h.highlight_venue_name][h.date] = h

        sorted_venues = sorted(venue_shows.keys())
        
        # Start footer 1 row below max_grid_row (User requested moving up by 1)
        footer_start_row = max_grid_row + 1
        current_row = footer_start_row
        
        # Styles for footer
        footer_venue_font = Font(name='Arial', size=11, bold=True, color='FF000000') # Black
        footer_time_font = Font(name='Arial', size=10, bold=True, color='FF000000') # Black
        footer_title_font = Font(name='Arial', size=11, bold=True, color='FF000000') # Black
        
        medium_side = Side(style='medium')
        thin_side = Side(style='thin')

        for venue in sorted_venues:
            # Row A (Time) and Row B (Title)
            row_a = current_row
            row_b = current_row + 1
            last_col = len(itinerary_items) + 2
            
            # Set row height to 25px (approx 18.75 points)
            ws.row_dimensions[row_a].height = 18.75
            ws.row_dimensions[row_b].height = 18.75
            
            # Venue Name (Left Col)
            ws.merge_cells(start_row=row_a, start_column=1, end_row=row_b, end_column=1)
            cell_left = ws.cell(row=row_a, column=1, value=venue)
            cell_left.font = footer_venue_font
            cell_left.alignment = center_align
            
            # Venue Name (Right Col)
            ws.merge_cells(start_row=row_a, start_column=last_col, end_row=row_b, end_column=last_col)
            cell_right = ws.cell(row=row_a, column=last_col, value=venue)
            cell_right.font = footer_venue_font
            cell_right.alignment = center_align

            # Shows per day
            for idx, item in enumerate(itinerary_items):
                col_idx = idx + 2
                show = venue_shows[venue].get(item.date)
                
                # Cell A (Time)
                cell_a = ws.cell(row=row_a, column=col_idx)
                cell_a.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
                
                # Cell B (Title)
                cell_b = ws.cell(row=row_b, column=col_idx)
                cell_b.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
                
                if show:
                    cell_a.value = show.time_text
                    cell_a.font = footer_time_font
                    
                    cell_b.value = show.title
                    cell_b.font = footer_title_font
            
            # Apply Borders to the entire 2-row block for this venue
            # We need a medium border around the outside (Top of row_a, Bottom of row_b, Left of col 1, Right of last_col)
            # And medium vertical borders for all columns to match the grid above.
            
            for c in range(1, last_col + 1):
                # Row A (Top)
                cell_top = ws.cell(row=row_a, column=c)
                current_top_border = cell_top.border
                new_top = medium_side
                new_bottom = Side(style=None) # No internal horizontal divider
                new_left = medium_side # Vertical lines are medium
                new_right = medium_side # Vertical lines are medium
                
                cell_top.border = Border(top=new_top, bottom=new_bottom, left=new_left, right=new_right)
                cell_top.fill = white_fill

                # Row B (Bottom)
                cell_bottom = ws.cell(row=row_b, column=c)
                current_bottom_border = cell_bottom.border
                new_top_b = Side(style=None) # No internal horizontal divider
                new_bottom_b = medium_side
                new_left_b = medium_side # Vertical lines are medium
                new_right_b = medium_side # Vertical lines are medium
                
                cell_bottom.border = Border(top=new_top_b, bottom=new_bottom_b, left=new_left_b, right=new_right_b)
                cell_bottom.fill = white_fill
            
            current_row += 2

    # Output
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{ship_code} {venue_name} Schedule - VY{voyage_number} - {datetime.now().strftime('%Y.%m.%d')}.xlsx"
    
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
